import cv2
import pytesseract
import re
from pytesseract import Output
import os
import openpyxl
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def extract_text_from_image(image_path, template_path):
    image = cv2.imread(image_path)
    image_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # применяем пороговую бинаризацию
    _, image_thresholded = cv2.threshold(image_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    template = cv2.imread(template_path, 0)

    w, h = template.shape[::-1]

    res = cv2.matchTemplate(image_thresholded, template, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)

    top_left = max_loc
    bottom_right = (top_left[0] + w, image.shape[0])

    crop_img = image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]

    extracted_text = pytesseract.image_to_string(crop_img, lang="rus+eng", config="--psm 6")  # Измените аргумент lang на "rus+eng"

    return extracted_text.strip()

def write_data_to_excel(txt_file, excel_file):
    # Создаем новую книгу и добавляем лист
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Устанавливаем имена столбцов
    column_names = ["Фамилия", "Убийства", "Смерти", "КД"]
    for col, name in enumerate(column_names, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = name

    # Открываем текстовый файл и считываем данные
    with open(txt_file, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Записываем данные в ячейки Excel
    for row, line in enumerate(lines, start=2):  # начинаем с 2-й строки, так как первая строка содержит имена столбцов
        data = line.strip().split(" - ")
        for col, value in enumerate(data, start=1):
            cell = sheet.cell(row=row, column=col)
            cell.value = value

    # Сохраняем книгу в файл
    workbook.save(excel_file)




def main():
    image_path = "cd.png"
    templates = ["1.png", "2.png", "3.png"]
    stats = []

    for template in templates:
        text = extract_text_from_image(image_path, template)
        stats.append(text)

    surnames = list(filter(None, stats[0].split('\n')))
    kills = list(filter(None, stats[1].split('\n')))
    deaths = list(filter(None, stats[2].split('\n')))

    surnames = [re.sub(r'\(.*?\)', '', name).strip() for name in surnames]

    if len(surnames) == len(kills) == len(deaths):
        stats_file = "stats.txt"

        # Читаем текущие данные из файла
        if os.path.exists(stats_file):
            with open(stats_file, "r", encoding="utf-8") as f:
                current_data = [line.strip().split(" - ") for line in f.readlines()]
        else:
            current_data = []

        # Создаем словарь для хранения текущих данных
        stats_dict = {entry[0]: [int(entry[1]), int(entry[2])] for entry in current_data}

        # Обновляем статистику в словаре
        for i in range(len(surnames)):
            if "Фамилия - 2 - 2" not in f"{surnames[i]} - {kills[i]} - {deaths[i]}" and kills[i].isdigit() and deaths[
                i].isdigit():
                current_kills = int(kills[i])
                current_deaths = int(deaths[i])
                if surnames[i] in stats_dict:
                    stats_dict[surnames[i]][0] += current_kills
                    stats_dict[surnames[i]][1] += current_deaths
                else:
                    stats_dict[surnames[i]] = [current_kills, current_deaths]

        # Записываем обновленную статистику обратно в файл
        with open(stats_file, "w", encoding="utf-8") as f:
            for name, stats in stats_dict.items():
                kills = int(stats[0])
                deaths = int(stats[1])
                if deaths == 0:
                    kd = kills
                else:
                    kd = round(kills / deaths, 2)
                f.write(f"{name} - {stats[0]} - {stats[1]} - {kd}\n")



    else:

        print("Ошибка: Количество фамилий, убийств и смертей не совпадает.")

        print(f"Количество фамилий: {len(surnames)}")

        print(f"Количество убийств: {len(kills)}")

        print(f"Количество смертей: {len(deaths)}")

        print("\nСписок фамилий:")

        for surname in surnames:
            print(surname)

        print("\nСписок убийств:")

        for kill in kills:
            print(kill)

        print("\nСписок смертей:")

        for death in deaths:
            print(death)
    # После записи обновленной статистики обратно в файл, вызовите функцию write_data_to_excel:
    txt_file = "stats.txt"
    excel_file = "stats.xlsx"
    write_data_to_excel(txt_file, excel_file)


if __name__ == "__main__":
    main()